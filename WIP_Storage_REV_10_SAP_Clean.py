# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Main5.ui'
#
# Created by: PyQt5 UI code generator 5.13.2
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QGridLayout, QSizePolicy, QProgressDialog
from openpyxl import load_workbook
import xlsxwriter
from WIP_Rack_Position import Rack_position
from WIP_Rack_Position_list import Rack_position_list
from Rack_group_position_set import Rack_group_pos as Rack_group_position
import datetime
import shutil
import os
import sqlite3
import pandas as pd
import json
from Read_SAP_Excel import get_barcode_data,get_card_info,get_card_info_for_update
from qtwidgets import PasswordEdit
import time
import traceback, sys


# Multithreading Class
class WorkerSignals(QObject):
    '''
    Defines the signals available from a running worker thread.

    Supported signals are:

    finished
        No data
    
    error
        `tuple` (exctype, value, traceback.format_exc() )
    
    result
        `object` data returned from processing, anything

    progress
        `int` indicating % progress 

    '''
    finished = pyqtSignal()
    error = pyqtSignal(tuple)
    result = pyqtSignal(object)
    progress = pyqtSignal(int)

class Worker(QRunnable):
    '''
    Worker thread

    Inherits from QRunnable to handler worker thread setup, signals and wrap-up.

    :param callback: The function callback to run on this worker thread. Supplied args and 
                     kwargs will be passed through to the runner.
    :type callback: function
    :param args: Arguments to pass to the callback function
    :param kwargs: Keywords to pass to the callback function

    '''

    def __init__(self, fn, *args, **kwargs):
        super(Worker, self).__init__()

        # Store constructor arguments (re-used for processing)
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()    

        # Add the callback to our kwargs
        #self.kwargs['progress_callback'] = self.signals.progress        

    @pyqtSlot()
    def run(self):
        '''
        Initialise the runner function with passed args, kwargs.
        '''
        
        # Retrieve args/kwargs here; and fire processing using them
        try:
            result = self.fn(*self.args, **self.kwargs)
        except:
            traceback.print_exc()
            exctype, value = sys.exc_info()[:2]
            self.signals.error.emit((exctype, value, traceback.format_exc()))
        else:
            self.signals.result.emit(result)  # Return the result of the processing
        finally:
            self.signals.finished.emit()  # Done
        
class Ui_MainPage(object):

# GUI part --------------------------------------------------------------------------------------------------
# ------------------------------------ GUI Structure Part----------------------------------------------------
    def setupUi(self, MainPage):
        # Contruct Main Page as Stacked Widget
        MainPage.setObjectName("MainPage")
        MainPage.resize(844, 614)
        MainPage.setMinimumSize(QtCore.QSize(800, 600))
        self.centralwidget = QtWidgets.QWidget(MainPage)
        self.centralwidget.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.centralwidget.sizePolicy().hasHeightForWidth())
        self.centralwidget.setSizePolicy(sizePolicy)
        self.centralwidget.setMinimumSize(QtCore.QSize(810, 600))
        self.centralwidget.setMaximumSize(QtCore.QSize(1920, 1080))
        self.centralwidget.setObjectName("centralwidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.centralwidget)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.stackedWidgetWIP = QtWidgets.QStackedWidget(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.stackedWidgetWIP.sizePolicy().hasHeightForWidth())
        self.stackedWidgetWIP.setSizePolicy(sizePolicy)
        self.stackedWidgetWIP.setMinimumSize(QtCore.QSize(100, 145))
        self.stackedWidgetWIP.setMaximumSize(QtCore.QSize(1564, 1124))
        self.stackedWidgetWIP.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.stackedWidgetWIP.setObjectName("stackedWidgetWIP")

        # Create "Main Page" as  StackWidget page

        self.Main1 = QtWidgets.QWidget()
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.Main1.sizePolicy().hasHeightForWidth())
        self.Main1.setSizePolicy(sizePolicy)
        self.Main1.setObjectName("Main1")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.Main1)
        self.gridLayout_2.setObjectName("gridLayout_2")

        # Add Label 
        self.Load_Barcode_label = QtWidgets.QLabel(self.Main1)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.Load_Barcode_label.setFont(font)
        self.Load_Barcode_label.setObjectName("Load_Barcode_label")
        self.gridLayout_2.addWidget(self.Load_Barcode_label, 1, 0, 1, 1)


        self.UpdatedPath = QtWidgets.QLineEdit(self.Main1)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.UpdatedPath.sizePolicy().hasHeightForWidth())
        self.UpdatedPath.setSizePolicy(sizePolicy)
        self.UpdatedPath.setObjectName("UpdatedPath")
        self.gridLayout_2.addWidget(self.UpdatedPath, 1, 2, 1, 1)


        self.Excel_SAP_label = QtWidgets.QLabel(self.Main1)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.Excel_SAP_label.setFont(font)
        self.Excel_SAP_label.setObjectName("Excel_SAP_label")
        self.gridLayout_2.addWidget(self.Excel_SAP_label, 2, 0, 1, 1)


        self.SAP_Excel = QtWidgets.QLineEdit(self.Main1)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.SAP_Excel.sizePolicy().hasHeightForWidth())
        self.SAP_Excel.setSizePolicy(sizePolicy)
        self.SAP_Excel.setObjectName("SAP_Excel")
        self.gridLayout_2.addWidget(self.SAP_Excel, 2, 2, 1, 1)
        

        self.labeWIPl = QtWidgets.QLabel(self.Main1)
        self.labeWIPl.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.labeWIPl.sizePolicy().hasHeightForWidth())
        self.labeWIPl.setSizePolicy(sizePolicy)
        self.labeWIPl.setMinimumSize(QtCore.QSize(500, 200))
        font = QtGui.QFont()
        font.setFamily("Browallia New")
        font.setPointSize(36)
        font.setBold(True)
        font.setItalic(False)
        font.setUnderline(False)
        font.setWeight(75)
        font.setStrikeOut(False)
        font.setKerning(True)
        font.setStyleStrategy(QtGui.QFont.NoAntialias)
        self.labeWIPl.setFont(font)
        self.labeWIPl.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.labeWIPl.setAutoFillBackground(False)
        self.labeWIPl.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.labeWIPl.setFrameShadow(QtWidgets.QFrame.Raised)
        self.labeWIPl.setAlignment(QtCore.Qt.AlignCenter)
        self.labeWIPl.setObjectName("labeWIPl")
        self.gridLayout_2.addWidget(self.labeWIPl, 0, 0, 1, 5)
        spacerItem = QtWidgets.QSpacerItem(
            20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_2.addItem(spacerItem, 7, 2, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(
            20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        self.gridLayout_2.addItem(spacerItem1, 2, 2, 1, 1)
        self.UpdatWIPStorageBtn = QtWidgets.QPushButton(self.Main1)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.UpdatWIPStorageBtn.sizePolicy().hasHeightForWidth())
        self.UpdatWIPStorageBtn.setSizePolicy(sizePolicy)
        self.UpdatWIPStorageBtn.setMinimumSize(QtCore.QSize(331, 71))
        self.UpdatWIPStorageBtn.setMaximumSize(QtCore.QSize(4000, 100))
        self.UpdatWIPStorageBtn.setSizeIncrement(QtCore.QSize(0, 0))
        font = QtGui.QFont()
        font.setFamily("System")
        font.setPointSize(8)
        font.setBold(True)
        font.setWeight(75)
        self.UpdatWIPStorageBtn.setFont(font)
        self.UpdatWIPStorageBtn.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.UpdatWIPStorageBtn.setObjectName("UpdatWIPStorageBtn")
        self.gridLayout_2.addWidget(self.UpdatWIPStorageBtn, 3, 2, 1, 1)
        spacerItem2 = QtWidgets.QSpacerItem(
            100, 20, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem2, 3, 3, 1, 1)
        spacerItem3 = QtWidgets.QSpacerItem(
            40, 20, QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem3, 1, 4, 1, 1)
        self.BrowseButton = QtWidgets.QPushButton(self.Main1)
        self.BrowseButton.setObjectName("BrowseButton")
        self.gridLayout_2.addWidget(self.BrowseButton, 1, 3, 1, 1)

        self.BrowseSAPExcelButton = QtWidgets.QPushButton(self.Main1)
        self.BrowseSAPExcelButton.setObjectName("BrowseSAPButton")
        self.gridLayout_2.addWidget(self.BrowseSAPExcelButton, 2, 3, 1, 1)
        spacerItem4 = QtWidgets.QSpacerItem(
            40, 20, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem4, 3, 0, 1, 1)
        self.Exportlog = QtWidgets.QPushButton(self.Main1)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.Exportlog.sizePolicy().hasHeightForWidth())
        self.Exportlog.setSizePolicy(sizePolicy)
        self.Exportlog.setMinimumSize(QtCore.QSize(130, 45))
        self.Exportlog.setMaximumSize(QtCore.QSize(200, 90))
        self.Exportlog.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.Exportlog.setObjectName("Exportlog")
        self.gridLayout_2.addWidget(self.Exportlog, 9, 0, 1, 1)

        self.ExportExcelButton = QtWidgets.QPushButton(self.Main1)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.ExportExcelButton.sizePolicy().hasHeightForWidth())
        self.ExportExcelButton.setSizePolicy(sizePolicy)
        self.ExportExcelButton.setMinimumSize(QtCore.QSize(130, 45))
        self.ExportExcelButton.setMaximumSize(QtCore.QSize(200, 90))
        self.ExportExcelButton.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.ExportExcelButton.setObjectName("ExportExcelButton")
        self.gridLayout_2.addWidget(self.ExportExcelButton, 8, 0, 1, 1)

        self.ExportInputlog = QtWidgets.QPushButton(self.Main1)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.ExportInputlog.sizePolicy().hasHeightForWidth())
        self.ExportInputlog.setSizePolicy(sizePolicy)
        self.ExportInputlog.setMinimumSize(QtCore.QSize(130, 45))
        self.ExportInputlog.setMaximumSize(QtCore.QSize(200, 90))
        self.ExportInputlog.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.ExportInputlog.setObjectName("Exportlog")
        self.gridLayout_2.addWidget(self.ExportInputlog, 10, 0, 1, 1)

        self.Setting_btn = QtWidgets.QPushButton(self.Main1)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.Setting_btn.sizePolicy().hasHeightForWidth())
        self.Setting_btn.setSizePolicy(sizePolicy)
        self.Setting_btn.setMinimumSize(QtCore.QSize(100, 45))
        self.Setting_btn.setObjectName("Setting_btn")
        self.gridLayout_2.addWidget(self.Setting_btn, 10, 3, 1, 1)


        self.Reupdate_btn = QtWidgets.QPushButton(self.Main1)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.Reupdate_btn.sizePolicy().hasHeightForWidth())
        self.Reupdate_btn.setSizePolicy(sizePolicy)
        self.Reupdate_btn.setMinimumSize(QtCore.QSize(100, 45))
        self.Reupdate_btn.setObjectName("Setting_btn")
        self.gridLayout_2.addWidget(self.Reupdate_btn, 8, 3, 1, 1)

        self.Database_pull = QtWidgets.QPushButton(self.Main1)
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(
            self.Database_pull.sizePolicy().hasHeightForWidth())
        self.Database_pull.setSizePolicy(sizePolicy)
        self.Database_pull.setMinimumSize(QtCore.QSize(100, 45))
        self.Database_pull.setObjectName("Setting_btn")
        self.gridLayout_2.addWidget(self.Database_pull, 9, 3, 1, 1)

        # Add Main page to Stack Widget
        self.stackedWidgetWIP.addWidget(self.Main1)


        # Create "Setting Page" as Stack Widget

        self.Setting = QtWidgets.QWidget()
        self.Setting.setObjectName("Setting")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.Setting)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.Max_digit_set = QtWidgets.QPushButton(self.Setting)
        self.Max_digit_set.setObjectName("Max_digit_set")
        self.gridLayout_3.addWidget(self.Max_digit_set, 2, 4, 1, 1)


        # Make Query Button for Query from MRP System
            # Make as blank line edit
            # Server part
        self.Query_Server = QtWidgets.QLineEdit(self.Setting)
        self.Query_Server.setObjectName("Query_Server")
        self.gridLayout_3.addWidget(self.Query_Server, 6, 2, 1, 1)
        self.Query_Server_Label = QtWidgets.QLabel(self.Setting)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.Query_Server_Label.setFont(font)
        self.Query_Server_Label.setObjectName("Query Server")
        self.gridLayout_3.addWidget(self.Query_Server_Label, 6, 0, 1, 1)
        
            # Database part
        self.Query_Database = QtWidgets.QLineEdit(self.Setting)
        self.Query_Database.setObjectName("Query_Server")
        self.gridLayout_3.addWidget(self.Query_Database, 7, 2, 1, 1)

        self.Query_Database_Label = QtWidgets.QLabel(self.Setting)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.Query_Database_Label.setFont(font)
        self.Query_Database_Label.setObjectName("Query Database")
        self.gridLayout_3.addWidget(self.Query_Database_Label, 7, 0, 1, 1)

            # Username part
        self.Query_Username = QtWidgets.QLineEdit(self.Setting)
        self.Query_Username.setObjectName("Query_Username")
        self.gridLayout_3.addWidget(self.Query_Username, 8, 2, 1, 1)

        self.Query_Username_Label = QtWidgets.QLabel(self.Setting)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.Query_Username_Label.setFont(font)
        self.Query_Username_Label.setObjectName("Query Username")
        self.gridLayout_3.addWidget(self.Query_Username_Label, 8, 0, 1, 1)

            # Password
        self.Query_Password = PasswordEdit()
        self.Query_Password.setObjectName("Query_Password")
        self.gridLayout_3.addWidget(self.Query_Password, 9, 2, 1, 1)
        self.Query_Password_Label = QtWidgets.QLabel(self.Setting)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.Query_Password_Label.setFont(font)
        self.Query_Password_Label.setObjectName("Query Password")
        self.gridLayout_3.addWidget(self.Query_Password_Label, 9, 0, 1, 1)

        self.Set_Query_Authen = QtWidgets.QPushButton(self.Setting)
        self.Set_Query_Authen.setObjectName("Set_Query_Authen")
        self.gridLayout_3.addWidget(self.Set_Query_Authen, 9, 3, 1, 1)
        

        # make set maxnumber as the spin line edit
        self.MaximumProducID = QtWidgets.QLabel(self.Setting)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.MaximumProducID.setFont(font)
        self.MaximumProducID.setObjectName("MaximumProducID")
        self.gridLayout_3.addWidget(self.MaximumProducID, 2, 0, 1, 2)

        self.Back_setting = QtWidgets.QPushButton(self.Setting)
        self.Back_setting.setObjectName("Back_setting")
        self.gridLayout_3.addWidget(self.Back_setting, 12, 4, 1, 1)

        self.SettingLabel = QtWidgets.QLabel(self.Setting)
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(22)
        font.setBold(True)
        font.setWeight(75)
        self.SettingLabel.setFont(font)
        self.SettingLabel.setAlignment(QtCore.Qt.AlignCenter)
        self.SettingLabel.setObjectName("SettingLabel")
        self.gridLayout_3.addWidget(self.SettingLabel, 0, 0, 1, 5)

        spacerItem6 = QtWidgets.QSpacerItem(
            20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_3.addItem(spacerItem6, 11, 2, 1, 1)

        self.Maxnumber_box = QtWidgets.QSpinBox(self.Setting)
        self.Maxnumber_box.setObjectName("Maxnumber_box")
        self.gridLayout_3.addWidget(self.Maxnumber_box, 2, 2, 1, 1)
        self.SaveSetting = QtWidgets.QPushButton(self.Setting)
        self.SaveSetting.setObjectName("SaveSetting")
        self.gridLayout_3.addWidget(self.SaveSetting, 11, 4, 1, 1)

        self.Export_Error_Log = QtWidgets.QPushButton(self.Setting)
        self.Export_Error_Log.setObjectName("Export_Error_Log")
        self.gridLayout_3.addWidget(self.Export_Error_Log, 10, 4, 1, 1)

        self.ResetDatalog = QtWidgets.QPushButton(self.Setting)
        self.ResetDatalog.setObjectName("ResetDatalog")
        self.gridLayout_3.addWidget(self.ResetDatalog, 8, 4, 1, 1)

        self.ResetInputlog = QtWidgets.QPushButton(self.Setting)
        self.ResetInputlog.setObjectName("ResetInputlog")
        self.gridLayout_3.addWidget(self.ResetInputlog, 9, 4, 1, 1)

        spacerItem7 = QtWidgets.QSpacerItem(
            20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_3.addItem(spacerItem7, 6, 2, 1, 1)

        self.ResetRack = QtWidgets.QPushButton(self.Setting)
        self.ResetRack.setObjectName("ResetRack")
        self.gridLayout_3.addWidget(self.ResetRack, 7, 4, 1, 1)

        self.stackedWidgetWIP.addWidget(self.Setting)
        self.horizontalLayout.addWidget(self.stackedWidgetWIP)
        MainPage.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainPage)
        self.statusbar.setObjectName("statusbar")
        MainPage.setStatusBar(self.statusbar)
        
        self.retranslateUi(MainPage)
        self.stackedWidgetWIP.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainPage)              
# -------------------------------------End of GUI Structure ------------------------------------------------
# -------------------------------------Class Global Variable Part-------------------------------------------
        # Multi Thread Object
        self.threadpool = QThreadPool()
        # Excel Path
        self.InputPath = None
        self.SAPPath = None
        self.excel_database_path = ""
        self.excel_database_path_temp = None

        self.sql_server =""
        self.sql_database=""
        self.sql_username =""
        self.sql_password = ""
        self.updated_data = []
        self.unload_product = []
        self.Re_update_product = []
        self.Re_update_temp_product = []
        self.None_product = []
    
        self.Updating_dialog = QMessageBox()
        self.product_barcode_to_query = ""
        self.None_temp_product = []
        self.create_WIP_Rack_Temporary_table()
        self.max_number = 9
        self.load_setting_value()
        self.Maxnumber_box.setValue(self.max_number)
        self.excel_database_file = self.excel_database_path.split('/')[-1]
        self.excel_database_path_vlookup = self.excel_database_path.replace(
            '/', "\\")
        self.excel_path_split = self.excel_database_path_vlookup.replace(
            self.excel_database_file, "")
        

        # Query Parameter
        self.Query_Server.setText('{}'.format(self.sql_server))
        self.Query_Database.setText('{}'.format(self.sql_database))
        self.Query_Username.setText('{}'.format(self.sql_username))
        self.Query_Password.setText('{}'.format(self.sql_password))




        # Initialize some excel table
        self.create_WIP_log_table()
        self.create_WIP_Input_log_table()
        self.create_WIP_Error_log_table()
# -------------------------------------End of Class Global Variable-----------------------------------------
# -------------------------------------GUI Button Definition Part ---------------------------------------------------
    def retranslateUi(self, MainPage):
        _translate = QtCore.QCoreApplication.translate
        MainPage.setWindowTitle(_translate("MainPage", "WIP WAREHOUSE"))
        self.labeWIPl.setText(_translate("MainPage", "WIP WAREHOUSE SYSTEM"))
        self.Load_Barcode_label.setText(_translate("MainPage", "Load_Barcode"))
        self.Excel_SAP_label.setText(_translate("MainPage", "Load_Excel_SAP"))
        self.UpdatWIPStorageBtn.setText(
            _translate("MainPage", "Update WIP Storage"))

        self.BrowseButton.setText(_translate("MainPage", "Browse_scanned_file"))

        self.BrowseSAPExcelButton.setText(_translate("MainPage", "Browse_SAP"))


        self.Exportlog.setText(_translate("MainPage", "Export\n"
                                          "Output Log Sheet"))
        self.ExportExcelButton.setText(_translate("MainPage", "Export \n"
                                                  "Rack Sheet"))
        self.ExportInputlog.setText(_translate("MainPage", "Export \n"
                                               "Input Log Sheet"))
        self.Setting_btn.setText(_translate("MainPage", "Setting"))
        self.Reupdate_btn.setText(_translate("MainPage","Re_update"))
        self.Database_pull.setText(_translate("MainPage","Database_Pull"))

        self.Max_digit_set.setText(_translate("MainPage", "Set Max digit"))
        self.Query_Server_Label.setText(_translate("MainPage", "Server"))
        self.Query_Database_Label.setText(_translate("MainPage", "Database"))
        self.Query_Username_Label.setText(_translate("MainPage", "Username"))
        self.Query_Password_Label.setText(_translate("MainPage", "Password"))



        self.MaximumProducID.setText(_translate(
            "MainPage", "Maximum digit of ProductID "))
        self.Back_setting.setText(_translate("MainPage", "Back"))

        self.SettingLabel.setText(_translate("MainPage", "Setting"))
        self.SaveSetting.setText(_translate("MainPage", " Save setting"))
        self.Export_Error_Log.setText(_translate("MainPage", " Export Error log"))
        self.ResetRack.setText(_translate("MainPage", "Reset Rack"))

        self.Set_Query_Authen.setText(_translate("MainPage", "Set Query info"))
        
        self.ResetDatalog.setText(_translate("MainPage", "Reset Datalog"))
        self.ResetInputlog.setText(_translate("MainPage", "Reset Input log"))
# ------------------------------------ End of GUI Button Definition -----------------------------------------------
# ------------------------------------ GUI button Connect to method part ---------------------------------------------------------
        self.BrowseButton.clicked.connect(
            lambda: self.Browse_file_update(self.UpdatedPath))

        self.BrowseSAPExcelButton.clicked.connect(
            lambda: self.Browse_file_SAP(self.SAP_Excel)
        )
        
        self.UpdatWIPStorageBtn.clicked.connect(lambda: self.Update_Handler())
        self.ExportExcelButton.clicked.connect(
            lambda: self.Export_Matching_workbook())
        self.Exportlog.clicked.connect(
            lambda: self.Export_log_sheet()
        )
        self.Setting_btn.clicked.connect(
            lambda: self.stackedWidgetWIP.setCurrentIndex(1)
        )
        self.Reupdate_btn.clicked.connect(
            lambda: self.rework_handler()
        )
        self.Database_pull.clicked.connect(
            lambda: self.Database_pull_handler()  
        )
        self.Back_setting.clicked.connect(
            lambda: self.stackedWidgetWIP.setCurrentIndex(0)
        )
        self.ResetRack.clicked.connect(
            lambda: self.Reset_Rack_SQL()
        )
        self.Max_digit_set.clicked.connect(
            lambda: self.set_max_number(self.Maxnumber_box.value())
        )
        self.Set_Query_Authen.clicked.connect(
            lambda : self.set_query_user(self.Query_Server.text(),self.Query_Database.text(),self.Query_Username.text(),self.Query_Password.text())
        )
        self.SaveSetting.clicked.connect(
            lambda: self.save_setting_Handler()
        )

        self.ResetDatalog.clicked.connect(
            lambda: self.Reset_datalog()
        )
        self.ResetInputlog.clicked.connect(
            lambda: self.Reset_data_input_log()
        )
        self.ExportInputlog.clicked.connect(
            lambda: self.Export_Input_log_sheet()
        )
        self.Export_Error_Log.clicked.connect(
            lambda: self.Export_Error_log_sheet()
        )
# ------------------------------------ End of button Connect to method -----------------------------------------------

# METHOD Part-----------------------------------------------------------------------
 # Browsing
  # ------------------------------------ Browse Method Part -----------------------------------------------------------------
    def Browse_file_update(self, lineshow):
        filename = QFileDialog.getOpenFileName(filter="*.xlsx")
        self.InputPath = filename[0]
        if self.InputPath != "" and self.InputPath != None:
            lineshow.setText('{}'.format(self.InputPath))
    
    def Browse_file_SAP(self, lineshow):
        filename = QFileDialog.getOpenFileName(filter="*.xlsx")
        self.SAPPath = filename[0]
        if self.SAPPath != "" and self.SAPPath != None:
            lineshow.setText('{}'.format(self.SAPPath))
  # ------------------------------------ End Browse Method -----------------------------------------------------------------

 # Getting Data 
  #------------------------------------- Query (Change after SAP launch as Read from Excel)---------------------------------
    def Query_Data(self):
        #### SAP ####
    
   
        product_data_lookup = get_barcode_data(self.SAPPath)

        if not product_data_lookup.empty:
            self.Update_data_to_pickle(product_data_lookup)
        
        return product_data_lookup



    def updated_data_to_query(self,data):
        to_query_barcode = [] 
        query_string = ""
        for i in data :
            to_query_barcode.append(i[2])

        for index,j in enumerate(to_query_barcode):
            
            if index == 0 :
                query_string = query_string+"'{}'".format(j)

            else:
                query_string = query_string + ",'{}'".format(j)
            
            #print(query_string)

        return(query_string)
    
    def get_None_product_to_query(self,None_product):
        
        to_query_barcode = []
        query_string = ""
        if None_product != [] :
            for i in None_product:
                to_query_barcode.append(i[1])
                #rework_data.append([2, i[0], i[1],i[2]])
            for index,j in enumerate(to_query_barcode):
                
                if index == 0 :
                    query_string = query_string+"'{}'".format(j)

                else:
                    query_string = query_string + ",'{}'".format(j)
                
                #print(query_string)

            return(query_string)
        else :
            return("")
  #------------------------------------- End Query -------------------------------------------------------------------------

 # SQLITE Database Interact
  #--------------------------------------- Error log database  ----------------------------------------
   # Error log
   # Create Error Database
    def create_WIP_Error_log_table(self):
        conn = sqlite3.connect('WIP_Storage.db')

        conn.execute(
            """CREATE TABLE IF NOT EXISTS ERROR_LOG (
         Error_Code         TEXT        NOT NULL,
         Description        TEXT,
         Time               Timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
         );"""
        )

        trigger_cmd = """
        CREATE TRIGGER IF NOT EXISTS ERROR_LOG AFTER INSERT ON ERROR_LOG
        BEGIN
            DELETE FROM ERROR_LOG where Error_Code NOT IN (SELECT Error_Code from ERROR_LOG ORDER BY Time DESC LIMIT 1000);
        END;
        """
        conn.execute(trigger_cmd)
        conn.close()
  # Put Error to Database
    def Put_Error(self, data):
        conn = sqlite3.connect('WIP_Storage.db')
        sql_cmd = """INSERT INTO ERROR_LOG(
            Error_Code,
            Description,
            Time
            ) VALUES ('{val1}','{val2}','{val3}')""".format(
            val1=data[0],
            val2=data[1],
            val3=data[2],
        )
        try:
            conn.execute(sql_cmd)
        except sqlite3.IntegrityError:
            #self.Show_duplicate_value_log(data[1])
            pass

        conn.commit()
        conn.close()

    def Error_001_log(self, data):
        error_code = "001"
        description = "Product : {} Not found".format(data)
        now = datetime.datetime.now()
        current_time = now.strftime(r"%Y-%m-%d %H:%M:%S")
        Time = current_time
        Error_packet = [error_code, description, Time]
        self.Put_Error(Error_packet)

    def Error_001_ext(self,data):
        error_code = "001_ext"
        description = "SQL Product Value : {} ".format(data)
        now = datetime.datetime.now()
        current_time = now.strftime(r"%Y-%m-%d %H:%M:%S")
        Time = current_time
        Error_packet = [error_code, description, Time]
        self.Put_Error(Error_packet)

    def Error_002_log(self, data):
        error_code = "002"
        description = "Rack : {} Not found".format(data)
        now = datetime.datetime.now()
        current_time = now.strftime(r"%Y-%m-%d %H:%M:%S")
        Time = current_time
        Error_packet = [error_code, description, Time]
        self.Put_Error(Error_packet)

    def Error_003_log(self, data):
        error_code = "003"
        description = "Wrong_input_file : {}".format(data)
        now = datetime.datetime.now()
        current_time = now.strftime(r"%Y-%m-%d %H:%M:%S")
        Time = current_time
        Error_packet = [error_code, description, Time]
        self.Put_Error(Error_packet)

    def Error_004_log(self, data):
        error_code = "004"
        description = "Peimission to : {} Error".format(data)
        now = datetime.datetime.now()
        current_time = now.strftime(r"%Y-%m-%d %H:%M:%S")
        Time = current_time
        Error_packet = [error_code, description, Time]
        self.Put_Error(Error_packet)

    def Warning_101_log(self, data):
        if data[1] != data[2]:
            error_code = '101_Incorrect Scan'
        else:
            error_code = "101"
            
        description = "Try Replace detect on {} _ pending replace {} by {} ".format(
            data[0], data[1], data[2])
        now = datetime.datetime.now()
        current_time = now.strftime(r"%Y-%m-%d %H:%M:%S")
        Time = current_time
        Error_packet = [error_code, description, Time]
        self.Put_Error(Error_packet)

    def Warning_102_log(self, data):
        error_code = "102"
        description = "Replace  on {} _ overwrite replace {} by {} ".format(
            data[0], data[1], data[2])
        now = datetime.datetime.now()
        current_time = now.strftime(r"%Y-%m-%d %H:%M:%S")
        Time = current_time
        Error_packet = [error_code, description, Time]
        self.Put_Error(Error_packet)

    def Warning_103_log(self, data):
        error_code = "103"
        description = "output product :{} is duplicated output".format(data)
        now = datetime.datetime.now()
        current_time = now.strftime(r"%Y-%m-%d %H:%M:%S")
        Time = current_time
        Error_packet = [error_code, description, Time]
        self.Put_Error(Error_packet)
  # Rack Data SQL 
   # Create Rack output table 
    def create_WIP_log_table(self):
        conn = sqlite3.connect('WIP_Storage.db')

        trigger_cmd = """
        CREATE TRIGGER IF NOT EXISTS WIP_log_trigger AFTER INSERT ON WIP_STORAGE_log
        BEGIN
            DELETE FROM WIP_STORAGE_log where Batch NOT IN (SELECT Batch from WIP_STORAGE_log ORDER BY Time DESC LIMIT 400000);
        END;
        """
        # temp.append([Rack_ID,productID,Glove,Glove_Path,Copyform,Cline,Cdate,online,weight,new_result,mi_result,mj_result,remark,aemployee,Size,Age_month_PF,Age_group_PF,DEPT,Product,Surface,Pro_Type,Length,Special,AGE,MPcs,PRONAME,Status,current_time])
        conn.execute(
            """CREATE TABLE IF NOT EXISTS WIP_STORAGE_log (
         RACK_ID            TEXT,
         Batch              TEXT    PRIMARY KEY     NOT NULL,
         GRTP               TEXT,
         SLOC               TEXT,
         Copyform           TEXT,
         Cline              TEXT,
         Cdate              TEXT,
         Date_QC_ทำรับ       TEXT,
         Weight             TEXT,
         QC_Total           TEXT,
         BR_AQL             TEXT,
         CR_AQL             TEXT,
         MJ_AQL             TEXT,
         MN_AQL             TEXT,
         PT_AQL             TEXT,
         Remark             TEXT,
         Remark2            TEXT,
         สถานะอายุ           TEXT,
         วันหมดอายุ_ครบ6เดือน             TEXT,
         ชิ้นพร้อมใช้UR                     TEXT,
         Blocked                        TEXT,
         รอตรวจสอบ                      TEXT,
         Status                         TEXT,
         Time               Timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
         );"""
        )

        conn.execute(trigger_cmd)
        conn.close()
   # Create Rack input table
    def create_WIP_Input_log_table(self):
        conn = sqlite3.connect('WIP_Storage.db')

        trigger_cmd = """
        CREATE TRIGGER IF NOT EXISTS WIP_log_trigger AFTER INSERT ON WIP_STORAGE_Input_log
        BEGIN
            DELETE FROM WIP_STORAGE_Input_log where Batch NOT IN (SELECT Batch from WIP_STORAGE_Input_log ORDER BY Time DESC LIMIT 400000);
        END;
        """
        # temp.append([Rack_ID,productID,Glove,Glove_Path,Copyform,Cline,Cdate,online,weight,new_result,mi_result,mj_result,remark,aemployee,Size,Age_month_PF,Age_group_PF,DEPT,Product,Surface,Pro_Type,Length,Special,AGE,MPcs,PRONAME,Status,current_time])
        conn.execute(
            """CREATE TABLE IF NOT EXISTS WIP_STORAGE_Input_log (
         RACK_ID            TEXT,
         Batch         TEXT    PRIMARY KEY     NOT NULL,
         GRTP              TEXT,
         SLOC               TEXT,
         Copyform           TEXT,
         Cline              TEXT,
         Cdate              TEXT,
         Date_QC_ทำรับ              TEXT,
         Weight             TEXT,
         QC_Total         TEXT,
         BR_AQL          TEXT,
         CR_AQL          TEXT,
         MJ_AQL             TEXT,
         MN_AQL          TEXT,
         PT_AQL               TEXT,
         Remark       REAL,
         Remark2       REAL,
         สถานะอายุ               TEXT,
        วันหมดอายุ_ครบ6เดือน            TEXT,
         ชิ้นพร้อมใช้UR            TEXT,
         Blocked           TEXT,
         รอตรวจสอบ             TEXT,
         Status             TEXT,
         Time               Timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
         );"""
        )

        conn.execute(trigger_cmd)
        conn.close()
   # Update data to output table Batch operate
    def update_to_output_log(self,data):
        conn = sqlite3.connect('WIP_Storage.db')
        for i in data:
            self.update_to_sql_log(i,conn)
        conn.commit()
        conn.close() 
   # Update data to output table SQLITE CMD
    def update_to_sql_log(self, data,conn):
        #conn = sqlite3.connect('WIP_Storage.db')
        # put data to sql log
        sql_cmd = """INSERT INTO WIP_STORAGE_log(
            RACK_ID,
            Batch,
            GRTP,
            SLOC,
            Copyform,
            Cline,
            Cdate,
            Date_QC_ทำรับ,
            Weight,
            QC_Total,
            BR_AQL,
            CR_AQL,
            MJ_AQL,
            MN_AQL,
            PT_AQL,
            Remark,
            Remark2,
            สถานะอายุ,
           วันหมดอายุ_ครบ6เดือน,
            ชิ้นพร้อมใช้UR,
            Blocked,
            รอตรวจสอบ,
            Status,
            Time
            ) VALUES ('{val1}','{val2}','{val3}','{val4}','{val5}','{val6}','{val7}','{val8}','{val9}','{val10}','{val11}','{val12}','{val13}','{val14}','{val15}','{val16}','{val17}','{val18}','{val19}','{val20}','{val21}','{val22}','{val23}','{val24}')""".format(
            val1=data[0],
            val2=data[1],
            val3=data[2],
            val4=data[3],
            val5=data[4],
            val6=data[5],
            val7=data[6],
            val8=data[7],
            val9=data[8],
            val10=data[9],
            val11=data[10],
            val12=data[11],
            val13=data[12],
            val14=data[13],
            val15=data[14],
            val16=data[15],
            val17=data[16],
            val18=data[17],
            val19=data[18],
            val20=data[19],
            val21=data[20],
            val22=data[21],
            val23=data[22],
            val24=data[23]

        )
        try:
            conn.execute(sql_cmd)
        except sqlite3.IntegrityError:
            pass
            #self.Warning_103_log(data[1])
            #self.Show_duplicate_value_log(data[1])

        #conn.commit()
        #conn.close()
   # Update data to input table Batch operate
    def update_to_Input_log(self, data):
        conn = sqlite3.connect('WIP_Storage.db')
        for i in data :
            self.update_to_sql_Input_log(i,conn)
        conn.commit()
        conn.close()
   # Update data to input table SQLITE CMD
    def update_to_sql_Input_log(self, data,conn):
        #conn = sqlite3.connect('WIP_Storage.db')
        # put data to sql log
        sql_cmd = """INSERT INTO WIP_STORAGE_Input_log(
            RACK_ID,
            Batch,
            GRTP,
            SLOC,
            Copyform,
            Cline,
            Cdate,
            Date_QC_ทำรับ,
            Weight,
            QC_Total,
            BR_AQL,
            CR_AQL,
            MJ_AQL,
            MN_AQL,
            PT_AQL,
            Remark,
            Remark2,
            สถานะอายุ,
           วันหมดอายุ_ครบ6เดือน,
            ชิ้นพร้อมใช้UR,
            Blocked,
            รอตรวจสอบ,
            Status,
            Time
            ) VALUES ('{val1}','{val2}','{val3}','{val4}','{val5}','{val6}','{val7}','{val8}','{val9}','{val10}','{val11}','{val12}','{val13}','{val14}','{val15}','{val16}','{val17}','{val18}','{val19}','{val20}','{val21}','{val22}','{val23}','{val24}')""".format(
            val1=data[0],
            val2=data[1],
            val3=data[2],
            val4=data[3],
            val5=data[4],
            val6=data[5],
            val7=data[6],
            val8=data[7],
            val9=data[8],
            val10=data[9],
            val11=data[10],
            val12=data[11],
            val13=data[12],
            val14=data[13],
            val15=data[14],
            val16=data[15],
            val17=data[16],
            val18=data[17],
            val19=data[18],
            val20=data[19],
            val21=data[20],
            val22=data[21],
            val23=data[22],
            val24=data[23]
           
        )
        try:
            conn.execute(sql_cmd)
        except sqlite3.IntegrityError:
            #self.Show_duplicate_value_log(data[1])
            pass
        #conn.commit()
        #conn.close()
   # Update old product data Batch
    def update_old_data_temp_Rack(self,data):
        conn = sqlite3.connect('WIP_Storage.db')
        for i in data :
            self.update_old_sql_temp_Rack(i,conn)
        conn.commit()
        conn.close() 
   # Update old product data Batch SQLite CMD
    def update_old_sql_temp_Rack(self, data,conn):
        #conn = sqlite3.connect('WIP_Storage.db')
        # put data to sql log
        sql_cmd = """UPDATE WIP_STORAGE_TEMP_Rack 
        SET  
            Batch = '{val2}',
            GRTP = '{val3}',
            SLOC = '{val4}',
            Copyform = '{val5}',
            Cline = '{val6}',
            Cdate = '{val7}',
            Date_QC_ทำรับ = '{val8}',
            Weight = '{val9}',
            QC_Total = '{val10}',
            BR_AQL = '{val11}',
            CR_AQL = '{val12}',
            MJ_AQL = '{val13}',
            MN_AQL = '{val14}',
            PT_AQL = '{val15}',
            Remark = '{val16}',
            Remark2 = '{val17}',
            สถานะอายุ = '{val18}',
            วันหมดอายุ_ครบ6เดือน = '{val19}',
            ชิ้นพร้อมใช้UR = '{val20}',
            Blocked= '{val21}',
            รอตรวจสอบ= '{val22}',
            Status = '{val23}',
            Time = '{val24}'
        WHERE 
            RACK_ID = '{val1}' """.format(
            val1=data[0],
            val2=data[1],
            val3=data[2],
            val4=data[3],
            val5=data[4],
            val6=data[5],
            val7=data[6],
            val8=data[7],
            val9=data[8],
            val10=data[9],
            val11=data[10],
            val12=data[11],
            val13=data[12],
            val14=data[13],
            val15=data[14],
            val16=data[15],
            val17=data[16],
            val18=data[17],
            val19=data[18],
            val20=data[19],
            val21=data[20],
            val22=data[21],
            val23=data[22],
            val24=data[23]
        ) 
        try:
            conn.execute(sql_cmd)
        except sqlite3.IntegrityError:
            
            pass
   # Update Data to temp Rack Batch operation
    def update_to_temp_Rack(self, data):
        conn = sqlite3.connect('WIP_Storage.db')
        for i in data :
            self.update_to_sql_temp_Rack(i,conn)
        conn.commit()
        conn.close() 
   # Update Data to temp Rack CMD
    def update_to_sql_temp_Rack(self, data,conn):
        #conn = sqlite3.connect('WIP_Storage.db')
        # put data to sql log
        sql_cmd = """INSERT INTO WIP_STORAGE_TEMP_Rack(
            RACK_ID,
            Batch,
            GRTP,
            SLOC,
            Copyform,
            Cline,
            Cdate,
            Date_QC_ทำรับ,
            Weight,
            QC_Total,
            BR_AQL,
            CR_AQL,
            MJ_AQL,
            MN_AQL,
            PT_AQL,
            Remark,
            Remark2,
            สถานะอายุ,
            วันหมดอายุ_ครบ6เดือน,
            ชิ้นพร้อมใช้UR,
            Blocked,
            รอตรวจสอบ,
            Status,
            Time

            ) VALUES ('{val1}','{val2}','{val3}','{val4}','{val5}','{val6}','{val7}','{val8}','{val9}','{val10}','{val11}','{val12}','{val13}','{val14}','{val15}','{val16}','{val17}','{val18}','{val19}','{val20}','{val21}','{val22}','{val23}','{val24}')""".format(
            val1=data[0],
            val2=data[1],
            val3=data[2],
            val4=data[3],
            val5=data[4],
            val6=data[5],
            val7=data[6],
            val8=data[7],
            val9=data[8],
            val10=data[9],
            val11=data[10],
            val12=data[11],
            val13=data[12],
            val14=data[13],
            val15=data[14],
            val16=data[15],
            val17=data[16],
            val18=data[17],
            val19=data[18],
            val20=data[19],
            val21=data[20],
            val22=data[21],
            val23=data[22],
            val24=data[23]

        )
        try:
            conn.execute(sql_cmd)
        except sqlite3.IntegrityError:
            #self.Show_duplicate_value_log(data[1])
            pass
   # Create WIP RACK Table
    def create_WIP_Rack_table(self):
        conn = sqlite3.connect('WIP_Storage.db')

        # temp.append([Rack_ID,productID,Glove,Glove_Path,Copyform,Cline,Cdate,online,weight,new_result,mi_result,mj_result,remark,aemployee,Size,Age_month_PF,Age_group_PF,DEPT,Product,Surface,Pro_Type,Length,Special,AGE,MPcs,PRONAME,Status,current_time])
        conn.execute(
            """CREATE TABLE IF NOT EXISTS WIP_STORAGE_Rack (
         RACK_ID            TEXT    PRIMARY KEY     NOT NULL,
         Batch         TEXT,    
         GRTP              TEXT,
         SLOC               TEXT,
         Copyform           TEXT,
         Cline              TEXT,
         Cdate              TEXT,
         Date_QC_ทำรับ              TEXT,
         Weight             TEXT,
         QC_Total         TEXT,
         BR_AQL          TEXT,
         CR_AQL          TEXT,
         MJ_AQL             TEXT,
         MN_AQL          TEXT,
         PT_AQL               TEXT,
         Remark       TEXT,
         Remark2       TEXT,
         สถานะอายุ               TEXT,
        วันหมดอายุ_ครบ6เดือน            TEXT,
         ชิ้นพร้อมใช้UR            TEXT,
         Blocked           TEXT,
         รอตรวจสอบ             TEXT,
         Status             TEXT,
         Time               Timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
         Updated_Time       Timestamp DATETIME
         );"""
        )

        for i in Rack_position_list:

            data = [i, "", "", "", "", "", "", "", "", "", "", "", "", "",
                    "", "", "", "", "", "", "", "", "", "", "", "", "", "",""]
            sql_add_rack_cmd = """INSERT INTO WIP_STORAGE_Rack(
                RACK_ID,
                Batch,
                GRTP,
                SLOC,
                Copyform,
                Cline,
                Cdate,
                Date_QC_ทำรับ,
                Weight,
                QC_Total,
                BR_AQL,
                CR_AQL,
                MJ_AQL,
                MN_AQL,
                PT_AQL,
                Remark,
                Remark2,
                สถานะอายุ,
               วันหมดอายุ_ครบ6เดือน,
                ชิ้นพร้อมใช้UR,
                Blocked,
                รอตรวจสอบ,
                Status,
                Time,
                Updated_Time
                ) VALUES ('{val1}','{val2}','{val3}','{val4}','{val5}','{val6}','{val7}','{val8}','{val9}','{val10}','{val11}','{val12}','{val13}','{val14}','{val15}','{val16}','{val17}','{val18}','{val19}','{val20}','{val21}','{val22}','{val23}','{val24}','{val25}')""".format(
                val1=data[0],
                val2=data[1],
                val3=data[2],
                val4=data[3],
                val5=data[4],
                val6=data[5],
                val7=data[6],
                val8=data[7],
                val9=data[8],
                val10=data[9],
                val11=data[10],
                val12=data[11],
                val13=data[12],
                val14=data[13],
                val15=data[14],
                val16=data[15],
                val17=data[16],
                val18=data[17],
                val19=data[18],
                val20=data[19],
                val21=data[20],
                val22=data[21],
                val23=data[22],
                val24=data[23],
                val25=data[24]
            )

            conn.execute(sql_add_rack_cmd)
            conn.commit()
        conn.close()
   # Create Temp Rack Table
    def create_WIP_Rack_Temporary_table(self):
        conn = sqlite3.connect('WIP_Storage.db')

        # temp.append([Rack_ID,productID,Glove,Glove_Path,Copyform,Cline,Cdate,online,weight,new_result,mi_result,mj_result,remark,aemployee,Size,Age_month_PF,Age_group_PF,DEPT,Product,Surface,Pro_Type,Length,Special,AGE,MPcs,PRONAME,Status,current_time])
        conn.execute(
            """CREATE TABLE IF NOT EXISTS WIP_STORAGE_TEMP_Rack (
         RACK_ID            TEXT    PRIMARY KEY     NOT NULL,
         Batch         TEXT,    
         GRTP              TEXT,
         SLOC               TEXT,
         Copyform           TEXT,
         Cline              TEXT,
         Cdate              TEXT,
         Date_QC_ทำรับ              TEXT,
         Weight             TEXT,
         QC_Total         TEXT,
         BR_AQL          TEXT,
         CR_AQL          TEXT,
         MJ_AQL             TEXT,
         MN_AQL          TEXT,
         PT_AQL               TEXT,
         Remark       REAL,
         Remark2       REAL,
         สถานะอายุ               TEXT,
        วันหมดอายุ_ครบ6เดือน            TEXT,
         ชิ้นพร้อมใช้UR            TEXT,
         Blocked           TEXT,
         รอตรวจสอบ             TEXT,
         Status             TEXT,
         Time               Timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
         Updated_Time       Timestamp DATETIME
         );"""
        )

        
        
        conn.commit()
        conn.close()
   
    def update_to_rack(self,data):
        conn = sqlite3.connect('WIP_Storage.db')
        
        for i in data:
            self.update_to_sql_Rack(i,conn)
        conn.commit()
        conn.close()

    def update_to_sql_Rack(self, data,conn):
        # conn = sqlite3.connect('WIP_Storage.db')
        # put data to sql log

        sql_cmd = """UPDATE WIP_STORAGE_Rack 
        SET  
            Batch = '{val2}',
            GRTP = '{val3}',
            SLOC = '{val4}',
            Copyform = '{val5}',
            Cline = '{val6}',
            Cdate = '{val7}',
            Date_QC_ทำรับ = '{val8}',
            Weight = '{val9}',
            QC_Total = '{val10}',
            BR_AQL = '{val11}',
            CR_AQL = '{val12}',
            MJ_AQL = '{val13}',
            MN_AQL = '{val14}',
            PT_AQL = '{val15}',
            Remark = '{val16}',
            Remark2 = '{val17}',
            สถานะอายุ = '{val18}',
            วันหมดอายุ_ครบ6เดือน = '{val19}',
            ชิ้นพร้อมใช้UR = '{val20}',
            Blocked= '{val21}',
            รอตรวจสอบ= '{val22}',
            Status = '{val23}',
            Time = '{val24}',
            Updated_Time = '{val25}'
        WHERE 
            RACK_ID = '{val1}' """.format(
            val1=data[0],
            val2=data[1],
            val3=data[2],
            val4=data[3],
            val5=data[4],
            val6=data[5],
            val7=data[6],
            val8=data[7],
            val9=data[8],
            val10=data[9],
            val11=data[10],
            val12=data[11],
            val13=data[12],
            val14=data[13],
            val15=data[14],
            val16=data[15],
            val17=data[16],
            val18=data[17],
            val19=data[18],
            val20=data[19],
            val21=data[20],
            val22=data[21],
            val23=data[22],
            val24=data[23],
            val25=data[24]

        )
        
        try:
            conn.execute(sql_cmd)
        except sqlite3.IntegrityError:
            #self.Show_duplicate_value_log(data[1])
            pass
        # conn.commit()
        # conn.close()

    def get_data_from_sql_Rack_by_product(self, product):
        conn = sqlite3.connect('WIP_Storage.db')

        cursor = conn.cursor()

        sql_cmd = """ SELECT * 
        FROM 
            WIP_STORAGE_Rack 
        WHERE 
            Batch = '{}'
        """.format(product)

        cursor.execute(sql_cmd)
        records = cursor.fetchall()
        cursor.close()

        if (conn):
            conn.close()

        return records

    def get_data_from_sql_by_Rack(self, Rack):
        conn = sqlite3.connect('WIP_Storage.db')

        cursor = conn.cursor()

        sql_cmd = """ SELECT * 
        FROM 
            WIP_STORAGE_Rack 
        WHERE 
            RACK_ID = '{}'
        """.format(Rack)

        cursor.execute(sql_cmd)
        records = cursor.fetchall()
        cursor.close()

        if (conn):
            conn.close()

        return records

    def check_if_rack_free(self, Rack):
        ProductID = self.get_data_from_sql_by_Rack(Rack)[0][1]

        if ProductID == "" or ProductID == None:
            return True
        else:
            return False

    def check_product_available_in_temp_rack(self,product):
        conn = sqlite3.connect('WIP_Storage.db')

        cursor = conn.cursor()

        sql_cmd = """ SELECT Batch 
        FROM 
            WIP_STORAGE_TEMP_Rack 
        WHERE 
            Batch = '{}'
        """.format(product)

        cursor.execute(sql_cmd)
        records = cursor.fetchall()
        cursor.close()

        if (conn):
            conn.close()

        if (records != None) and (records != "") and (records != []):
            return True

        else:
            #self.Error_001_ext(records)
            
            return False

    def delete_in_temp_rack(self,product):
        conn = sqlite3.connect('WIP_Storage.db')

        cursor = conn.cursor()

        sql_cmd = """ DELETE  
        FROM 
            WIP_STORAGE_TEMP_Rack 
        WHERE 
            Batch = '{}'
        """.format(product)

        cursor.execute(sql_cmd)
        cursor.close()
        conn.commit()

    def check_if_product_available(self, product):
        conn = sqlite3.connect('WIP_Storage.db')

        cursor = conn.cursor()

        sql_cmd = """ SELECT Batch 
        FROM 
            WIP_STORAGE_Rack 
        WHERE 
            Batch = '{}'
        """.format(product)

        cursor.execute(sql_cmd)
        records = cursor.fetchall()
        cursor.close()

        if (conn):
            conn.close()

        if (records != None) and (records != "") and (records != []):
            return True

        else:
            #self.Error_001_ext(records)
            
            return False

    def delete_data_from_Rack(self, data):
        conn = sqlite3.connect('WIP_Storage.db')
        # put data to sql log
        sql_cmd = """UPDATE WIP_STORAGE_Rack 
        SET  
            Batch = "",
            GRTP = "",
            SLOC = "",
            Copyform = "",
            Cline = "",
            Cdate = "",
            Date_QC_ทำรับ = "",
            Weight = "",
            QC_Total = "",
            BR_AQL = "",
            CR_AQL = "",
            MJ_AQL = "",
            MN_AQL = "",
            PT_AQL = "",
            Remark = "",
            Remark2 = "",
            สถานะอายุ = "",
           วันหมดอายุ_ครบ6เดือน = "",
            ชิ้นพร้อมใช้UR = "",
            Blocked = "",
            รอตรวจสอบ = "",
            Time = NULL
        WHERE 
            RACK_ID = '{val1}' """.format(
            val1=data
        )
        try:
            conn.execute(sql_cmd)
        except sqlite3.IntegrityError:
            #self.Show_duplicate_value_log(data[1])
            pass
        conn.commit()
        conn.close()

    def delete_product_Rack(self,data):
        conn = sqlite3.connect('WIP_Storage.db')
        for i in data :
            self.delete_data_from_Rack_By_product(i,conn)

        conn.commit()
        conn.close()

    def delete_data_from_Rack_By_product(self, product,conn):
        now = datetime.datetime.now()
        time = now.strftime(r"%m-%d-%Y %H:%M:%S")
        update_time = time
        #conn = sqlite3.connect('WIP_Storage.db')
        # put data to sql log
        sql_cmd = """UPDATE WIP_STORAGE_Rack 
        SET  
            Batch = "",
            GRTP = "",
            SLOC = "",
            Copyform = "",
            Cline = "",
            Cdate = "",
            Date_QC_ทำรับ = "",
            Weight = "",
            QC_Total = "",
            BR_AQL = "",
            CR_AQL = "",
            MJ_AQL = "",
            MN_AQL = "",
            PT_AQL = "",
            Remark = "",
            Remark2 = "",
            สถานะอายุ = "",
            วันหมดอายุ_ครบ6เดือน = "",
            ชิ้นพร้อมใช้UR = "",
            Blocked= "",
            รอตรวจสอบ= "",
            Status = "",
            Time = "",
            Updated_Time = '{val2}'
        WHERE 
            Batch = '{val1}' """.format(
            val1=product,
            val2 =update_time
        )
        try:
            conn.execute(sql_cmd)
        except sqlite3.IntegrityError:
            
            pass
            # self.Show_Input_ProductID_Notfound_dialog(product)

    def make_sql_to_excel_log(self):
        conn = sqlite3.connect('WIP_Storage.db')
        sql_string = 'select * from WIP_STORAGE_log'
        df = pd.read_sql(sql_string, conn)
        df.fillna("", inplace=True)
        conn.close()
        df['Batch'] = ("'" + df['Batch'])
        df['Batch'] = df['Batch'].astype('str')
        df.to_csv("log_temp.csv", index=False)

        df2 = pd.read_csv("log_temp.csv", dtype='str')
        df2.fillna(" ", inplace=True)
        # List of column name dictionaries
        
        
        # Create and propagate workbook
        workbook = xlsxwriter.Workbook('WIP_Storage_log.xlsx', options={
            'nan_inf_to_errors': True ,'strings_to_numbers': True})
        worksheet1 = workbook.add_worksheet()


        numbersformat = workbook.add_format({'num_format': '@'})
        headers = [{"header": i, 'format': numbersformat } for i in list(df2.columns)]

        worksheet1.add_table(0, 0, len(df2), len(
            df2.columns)-1, {"columns": headers , "data": df2.values.tolist(), 'style': 'Table Style Light 8'})
        workbook.close()


        
    def make_sql_to_excel_log_Rack(self):
        conn = sqlite3.connect('WIP_Storage.db')
        sql_string = 'select * from WIP_STORAGE_Rack'
        sql_string_temp_rack = 'select * from WIP_STORAGE_TEMP_Rack'
        df = pd.read_sql(sql_string, conn)
        df_temp_rack = pd.read_sql(sql_string_temp_rack, conn)
        df.replace(to_replace="",
                   value=" ")
        df.replace(to_replace="nan",
                   value="None")

        df_temp_rack.replace(to_replace="",
                   value=" ")
        df_temp_rack.replace(to_replace="nan",
                   value="None")
        # print(df)
        conn.close()

        df.loc[df['Batch'] != "",['Batch']] = ("'" + df.loc[df['Batch'] != "",['Batch']])
        #df['PRODUCT_ID'] = ("'" + df['PRODUCT_ID'])
        df['Batch'] = df['Batch'].astype('str')
        df.to_csv("Rack_Product_Match_sheet_temp.csv", index=False)

        df_temp_rack.loc[df_temp_rack['Batch'] != "",['Batch']] = ("'" + df_temp_rack.loc[df_temp_rack['Batch'] != "",['Batch']])
        #df['PRODUCT_ID'] = ("'" + df['PRODUCT_ID'])
        df_temp_rack['Batch'] = df_temp_rack['Batch'].astype('str')
        df_temp_rack.to_csv("Rack_Product_Match_sheet_V_Rack_temp.csv", index=False)

        df2 = pd.read_csv("Rack_Product_Match_sheet_temp.csv", dtype='str')
        # List of column name dictionaries
        df2.replace(to_replace="nan",
                   value="None")
        df2.fillna(" ", inplace=True)
        # print(df2)
        #headers = [{"header": i} for i in list(df2.columns)]
        df_temp_rack2 = pd.read_csv("Rack_Product_Match_sheet_V_Rack_temp.csv", dtype='str')
        # List of column name dictionaries
        df_temp_rack2.replace(to_replace="nan",
                   value="None")
        df_temp_rack2.fillna(" ", inplace=True)



        # Create and propagate workbook
        workbook = xlsxwriter.Workbook('Rack_Product_Match_sheet.xlsx', options={
            'nan_inf_to_errors': True , 'strings_to_numbers': True })
        
        
        worksheet1 = workbook.add_worksheet("Rack table")
        numbersformat = workbook.add_format({'num_format': '@'})
        headers = [{"header": i, 'format': numbersformat } for i in list(df2.columns)]

        
        worksheet1.add_table(0, 0, len(df2), len(
            df2.columns)-1, {"name": "Main_Rack","columns": headers, "data": df2.values.tolist(), 'style': 'Table Style Light 8'})
        



        worksheet2 = workbook.add_worksheet("Rack virtual table")
        numbersformat = workbook.add_format({'num_format': '@'})
        headers = [{"header": i, 'format': numbersformat } for i in list(df_temp_rack2.columns)]
        #a = df_temp_rack2.values.tolist()
        #b = list(df_temp_rack2.columns)
        #print(df_temp_rack2.values.tolist())
        
        worksheet2.add_table(0, 0, len(df_temp_rack2), len(
            df_temp_rack2.columns)-1, {"name": "Virtual_Rack","columns": headers, "data": df_temp_rack2.values.tolist(), 'style': 'Table Style Light 8'})
        

        Find_Product_pos_col = ["Product","Rack"]
        worksheet3 = workbook.add_worksheet("Find Product")
        numbersformat = workbook.add_format({'num_format': '0'})
        headers = [{"header": i, 'format': numbersformat } for i in Find_Product_pos_col]
        worksheet3.add_table(0, 0, 1, len(
            Find_Product_pos_col)-1, {"name": "Find_Product_position","data": [["'Y000005229","""=IFNA(INDEX(Main_Rack[RACK_ID],MATCH(A2,Main_Rack[Batch],0)),INDEX(Main_Rack[RACK_ID],MATCH("'"&A2,Main_Rack[Batch],0)))"""]], "columns": headers , "style": "Table Style Light 8"})
        #worksheet3.write_formula('A3', """=IFNA(INDEX(Main_Rack[RACK_ID],MATCH(A2,Main_Rack[Batch],0)),INDEX(Main_Rack[RACK_ID],MATCH("'"&A2,Main_Rack[Batch],0)))""")

        non_pair_df = self.get_non_pair_df()
        if not non_pair_df.empty:
            worksheet4 = workbook.add_worksheet("Rack ที่ไม่มีคู่")
            numbersformat = workbook.add_format({'num_format': '@'})
            headers = [{"header": i, 'format': numbersformat } for i in list(non_pair_df.columns)]
            worksheet4.add_table(0, 0, len(non_pair_df), len(
            non_pair_df.columns)-1, {"name": "NonPair_Rack","columns": headers, "data": non_pair_df.values.tolist(), 'style': 'Table Style Light 8'})






        workbook.close()

    def make_input_log(self):
        conn = sqlite3.connect('WIP_Storage.db')
        sql_string = 'select * from WIP_STORAGE_Input_log'
        df = pd.read_sql(sql_string, conn)
        df.fillna("None", inplace=True)
        conn.close()

        df['Batch'] = ("'" + df['Batch'])
        df['Batch'] = df['Batch'].astype('str')
        df.to_csv("Input_log_temp.csv", index=False)

        df2 = pd.read_csv("Input_log_temp.csv", dtype='str')
        df2.fillna("None", inplace=True)
        # List of column name dictionaries
        #headers = [{"header": i} for i in list(df2.columns)]
        # Create and propagate workbook
        workbook = xlsxwriter.Workbook('WIP_Storage_Input_log.xlsx', options={
            'nan_inf_to_errors': True ,'strings_to_numbers': True})
        worksheet1 = workbook.add_worksheet()

        numbersformat = workbook.add_format({'num_format': '@'})
        headers = [{"header": i, 'format': numbersformat } for i in list(df2.columns)]


        worksheet1.add_table(0, 0, len(df2), len(
            df2.columns)-1, {"columns": headers, "data": df2.values.tolist(), 'style': 'Table Style Light 8'})
        workbook.close()

    def make_Error_log(self):
        conn = sqlite3.connect('WIP_Storage.db')
        sql_string = 'select * from ERROR_LOG'
        df = pd.read_sql(sql_string, conn)
        df.fillna("", inplace=True)
        conn.close()

        df.to_csv("Error_log_temp.csv", index=False)

        df2 = pd.read_csv("Error_log_temp.csv", dtype='str')
        df2.fillna(" ", inplace=True)
        # List of column name dictionaries
        headers = [{"header": i} for i in list(df2.columns)]
        # Create and propagate workbook
        workbook = xlsxwriter.Workbook('Error_Log.xlsx', options={
            'nan_inf_to_errors': True})
        worksheet1 = workbook.add_worksheet()
        worksheet1.add_table(0, 0, len(df2), len(
            df2.columns)-1, {"columns": headers, "data": df2.values.tolist(), 'style': 'Table Style Light 8'})
        workbook.close()

    def delete_all_record_log(self):
        conn = sqlite3.connect('WIP_Storage.db')
        c = conn.cursor()
        # delete all rows from table
        c.execute('DELETE FROM WIP_STORAGE_log;')
        print('We have deleted', c.rowcount, 'records from the table.')
        # commit the changes to db
        conn.commit()
        # close the connection
        conn.close()

    def delete_all_input_log(self):
        conn = sqlite3.connect('WIP_Storage.db')
        c = conn.cursor()
        # delete all rows from table
        c.execute('DELETE FROM WIP_STORAGE_Input_log;')
        print('We have deleted', c.rowcount, 'records from the table.')
        # commit the changes to db
        conn.commit()
        # close the connection
        conn.close()

    def reset_SQL_Rack_table(self):
        conn = sqlite3.connect('WIP_Storage.db')
        c = conn.cursor()
        c.execute('DROP TABLE WIP_STORAGE_Rack')
        conn.commit()
        conn.close()
        self.create_WIP_Rack_table()

    def reset_SQL_Rack_table_del_all_current(self):
        conn = sqlite3.connect('WIP_Storage.db')
        # put data to sql log
        sql_cmd = """UPDATE WIP_STORAGE_Rack 
        SET  
            Batch = "",
            GRTP = "",
            SLOC = "",
            Copyform = "",
            Cline = "",
            Cdate = "",
            Date_QC_ทำรับ = "",
            Weight = "",
            QC_Total = "",
            BR_AQL = "",
            CR_AQL = "",
            MJ_AQL = "",
            MN_AQL = "",
            PT_AQL = "",
            Remark = "",
            Remark2 = "",
            สถานะอายุ = "",
           วันหมดอายุ_ครบ6เดือน = "",
            ชิ้นพร้อมใช้UR = "",
            Blocked= "",
            รอตรวจสอบ= "",
            Status = "",
            Time = ""
        WHERE 
            Batch != '' """
        try:
            conn.execute(sql_cmd)
        except sqlite3.IntegrityError:
            pass

        conn.commit()
        conn.close()
  # SAP Modification
    def get_blank_product_from_db(self):
        conn = sqlite3.connect('WIP_Storage.db')

        cursor = conn.cursor()

        sql_cmd = """ SELECT RACK_ID,Batch,Time
        FROM 
            WIP_STORAGE_Rack
        WHERE
            ((Batch IS NOT NULL) AND (Batch != "") AND (Batch != " ") )
        """

        cursor.execute(sql_cmd)

        records = cursor.fetchall()

        cursor.close()

        if (conn):
            conn.close()

        return records

    def get_blank_TEMP_product_from_db_(self):
        conn = sqlite3.connect('WIP_Storage.db')

        cursor = conn.cursor()

        sql_cmd = """ SELECT RACK_ID,Batch,Time
        FROM 
            WIP_STORAGE_TEMP_Rack  
        WHERE
            ((Batch IS NOT NULL) AND (Batch != "") AND (Batch != " ") )
        """

        cursor.execute(sql_cmd)

        records = cursor.fetchall()
        
        cursor.close()

        if (conn):
            conn.close()

        return records
    
    def get_nonpair_blank_from_db(self):
        conn = sqlite3.connect('WIP_Storage.db')

        cursor = conn.cursor()

        sql_cmd = """ SELECT RACK_ID
        FROM 
            WIP_STORAGE_Rack
        WHERE
            Batch == ""
        """

        cursor.execute(sql_cmd)

        records = cursor.fetchall()

        cursor.close()

        if (conn):
            conn.close()


        return records

 # Excel Interact
  # --------------------------------- Excel interact --------------------------------------

    def get_data(self, file):
        Data = []
        Input_Wb = load_workbook(filename=file)
        Input_sheet = Input_Wb.active
        # print(Input_sheet.max_column)
        if Input_sheet.max_column == 1:
            first_column = Input_sheet['A']
            for x in first_column:

                if x.value != None and x.value != "" and x.value != "None":
                    Data.append(x)

            Input_Wb.close()
            if Data != [] and Data[-1].value == None:
                del Data[-1]
        return Data

    def Load_unload_separator(self, Data):
        Packed_Data = []
        Is_Rack_ID = False
        last_data = []
        RackID = ""
        Rack_group_Ready = False
        Rack_Buttom_Ready = False
        Rack_Top_Ready = False
        load_package = []
        Bottom_pos =""
        Top_pos = ""


        if len(Data) == 1 and Data[0].value == None:
            return None

        for i in Data:
            '''
            Package first : 0 Load
                        1 Unload
            if load package will have size == 3
                    second : RackID
                    third  : ProductID
            if unload pacakge will have size == 2
                    Second : ProductID
            '''
        # Strip Data to eliminate space 
            i.value = str(i.value)
            i.value = i.value.strip()

            
        # i.value is current data
            # check if it is rack postion load data to Load slot
            if i.value in Rack_position:
                RackID = i.value
                Is_Rack_ID = True

        # if current data is Product ID Load and last data is Rack Position
        # data is for load
            elif i.value not in Rack_position and Is_Rack_ID == True:
                if len(str(i.value)) < self.max_number:
                    ProductID = "0"+str(i.value)
                else:
                    ProductID = str(i.value)

                Load_package = [0, RackID, ProductID]
                Packed_Data.append(Load_package)

                Is_Rack_ID = False


            # if the Barcode is in Group Barcode do auto grouping

            elif i.value in Rack_group_position :
                Bottom_pos = i.value + "B"
                Top_pos = i.value+"T"
                Rack_group_Ready = True
                Rack_Buttom_Ready = True
                Rack_Top_Ready = True
            

            elif ((i.value not in Rack_group_position ) and (i.value not in Rack_position) and (Rack_group_Ready == True)):
                if i.value != last_data:
                    
                    zero_amount = abs(self.max_number-len(str(i.value)))
                    zero_fill_string = "0"*zero_amount 

                    if len(str(i.value)) < self.max_number:
                        ProductID =  zero_fill_string+str(i.value)
                    else:
                        ProductID = str(i.value)

                    if(Rack_Buttom_Ready and Rack_Top_Ready):
                        load_package = [0, Top_pos, ProductID]
                        Rack_Top_Ready = False

                    elif(Rack_Top_Ready == False and Rack_Buttom_Ready ==True):

                        load_package = [0, Bottom_pos, ProductID]
                        Rack_Buttom_Ready = False

                    Packed_Data.append(load_package)
                    if(Rack_Buttom_Ready == False and Rack_Top_Ready == False):
                        Rack_group_Ready = False

        # if current data is Product ID Load and last data is not Rack Position
        # the data is for unload
            else:
                if i.value != last_data:
                    Is_Rack_ID = False
                    Rack_group_Ready = False
                    Rack_Buttom_Ready = False
                    Rack_Top_Ready = False
                    zero_amount = abs(self.max_number-len(str(i.value)))
                    zero_fill_string = "0"*zero_amount 


                    if len(str(i.value)) < self.max_number:
                        ProductID =  zero_fill_string+str(i.value)
                    else:
                        ProductID = str(i.value)
                    Unload_package = [1, "Unload", ProductID]
                    Packed_Data.append(Unload_package)

                    # Unload_data.append(i.value)
            

            last_data = i.value



        print(Packed_Data)

        return Packed_Data

    def Export_Matching_workbook(self):
        self.make_sql_to_excel_log_Rack()
        Source = 'Rack_Product_Match_sheet.xlsx'
        Backup_database_source = 'WIP_STORAGE.db'
        Destination = QFileDialog.getSaveFileName(filter="*.xlsx")[0]
        remove_backup_file_Destination = Destination.split('/')[-1]
        #print(remove_backup_file_Destination) 
        Backup_Destination = Destination.replace(remove_backup_file_Destination,'WIP_STORAGE.db')
        #print(Destination)
        if Destination != "" and Destination != None:
            try:
                shutil.copyfile(Source, Destination)
                shutil.copyfile(Backup_database_source,Backup_Destination)
                self.Show_Export_complete_dialog()
            except PermissionError:
                self.Error_004_log(Destination)
                self.Show_Permission_Error_dialog()
            except:
                self.Show_Unknow_Error_dialog()

    def Export_log_sheet(self):
        self.make_sql_to_excel_log()
        Source = 'WIP_Storage_log.xlsx'
        Destination = QFileDialog.getSaveFileName(filter="*.xlsx")[0]
        # print(Destination)
        if Destination != "" and Destination != None:
            try:
                shutil.copyfile(Source, Destination)
                self.Show_Export_complete_dialog()
            except PermissionError:
                self.Error_004_log(Destination)
                self.Show_Permission_Error_dialog()
            except:
                self.Show_Unknow_Error_dialog()

    def Export_Input_log_sheet(self):
        self.make_input_log()
        Source = 'WIP_Storage_Input_log.xlsx'
        Destination = QFileDialog.getSaveFileName(filter="*.xlsx")[0]
        # print(Destination)
        if Destination != "" and Destination != None:
            try:
                shutil.copyfile(Source, Destination)
                self.Show_Export_complete_dialog()
            except PermissionError:
                self.Error_004_log(Destination)
                self.Show_Permission_Error_dialog()
            except:
                self.Show_Unknow_Error_dialog()

    def Export_Error_log_sheet(self):
        self.make_Error_log()
        Source = 'Error_Log.xlsx'
        Destination = QFileDialog.getSaveFileName(filter="*.xlsx")[0]
        # print(Destination)
        if Destination != "" and Destination != None:
            try:
                shutil.copyfile(Source, Destination)
                self.Show_Export_complete_dialog()
            except PermissionError:
                self.Error_004_log(Destination)
                self.Show_Permission_Error_dialog()
            except:
                self.Show_Unknow_Error_dialog()

 # Back up Data Store 
  #--------------------------------------- Store Reserve Data -----------------------------------------
    def Database_pull_handler(self):
        refresh_Excel_data = self.Ask_for_Update_data_rack()
        
        if refresh_Excel_data:
        
            self.product_barcode_to_query = ""
            worker = Worker(self.Query_Data) # Any other args, kwargs are passed to the run function
            worker.signals.result.connect(self.Update_data_to_pickle_pull_database)

        # Execute
            self.threadpool.start(worker)
            self.Show_Please_Wait_Update_dialog()
            
        else:
            return None   

    def Update_data_to_pickle(self,Update_data):

        product_data_lookup = pd.read_pickle("backup_data_info.pkl")
        pickle_update_data = pd.concat([product_data_lookup, Update_data])
        pickle_update_data = pickle_update_data.tail(8000).reset_index(drop=True)

        #pickle_update_data.drop(pickle_update_data[pickle_update_data.score < 50].index, inplace=True)
        pickle_update_data.to_pickle("backup_data_info.pkl")

    def Update_data_to_pickle_pull_database(self,Update_data):
        self.Updating_dialog.done(0)
        if not Update_data.empty :
            product_data_lookup = pd.read_pickle("backup_data_info.pkl")
            pickle_update_data = pd.concat([product_data_lookup, Update_data])
            pickle_update_data.drop_duplicates(inplace= True)
            pickle_update_data = pickle_update_data.tail(5000).reset_index(drop=True)
            
            #pickle_update_data.drop(pickle_update_data[pickle_update_data.score < 50].index, inplace=True)
            pickle_update_data.to_pickle("backup_data_info.pkl")
            self.Show_Finish_Pull_data()
        else :
            self.Show_Input_Wrong_file_dialog()
            return None

 # Updating Rack Data 
  #------------------------------------- Update Data -----------------------------------------------------
    # Update Data to SQLITE DB
    # Main Function to update data 
    def Update_data(self,Updated_dataframe):
        
        '''This function will separate load and unload automatically'''
        if not Updated_dataframe.empty :
            # Define blank local variable
            #self.Updating_dialog.done(0)
            Error = False
            self.unload_product = []
            self.load_product = [] 
            stack_unload_list = [] 
            stack_load_list = [] 
            stack_update_list = []
            temp_Rack_load_list = []
            temp_Rack_unload_list = [] 
            duplicated_in_rack = []
            duplicated_in_temp_rack = []
            detailed_info= get_card_info(self.updated_data,Updated_dataframe)
            #self.Updating_dialog.done(0)
            for dat in detailed_info:

                # remove value that will be replace by reupdate function

                if self.None_product !=[]:
                        for i,inf in enumerate(self.None_product):
                            if inf[0] == dat[1]:
                                self.None_product.remove(self.None_product[i])
                                #print("equal")
                # if Load Type
                if dat[0] == 0:
                    # Load
                    temp_prod_avail = self.check_product_available_in_temp_rack(dat[2])
                    duplicated_prod = self.check_if_product_available(dat[2])

                    if temp_prod_avail:
                        #self.delete_in_temp_rack(dat[2])
                        duplicated_in_temp_rack.append(dat[2])


                    if(self.check_if_rack_free(dat[1]) != True) :

                        CurrentProductInfo = self.get_data_from_sql_by_Rack(dat[1])[0][:-1]

                        CurrentProduct = CurrentProductInfo[1] 

                        if CurrentProduct != dat[2]:
                            warning_packt = [dat[1], CurrentProduct, dat[2]]
                            self.Warning_101_log(warning_packt)
                            Overwrite = self.Ask_for_overwrite_load(
                                To_write_Product_ID=dat[2], Rack_ID=dat[1], Current_Product_ID=CurrentProduct)
                        #Overwrite = True
                            if(Overwrite):
                                if stack_load_list !=[]:
                                    if dat[1] in stack_load_list[:][0]:
                                        res = next(x for x, val in enumerate(stack_load_list[:][0]) if val == dat[1]) 
                                        #duplicate_load_list.append(stack_load_list[res])
                                        temp_Rack_load_list.append(stack_load_list[res])

                                if duplicated_prod:
                                    #self.delete_product_Rack([dat[2]])
                                    duplicated_in_rack.append(dat[2])

                                stack_load_list.append(dat[1:])
                                temp_Rack_load_list.append(CurrentProductInfo)
                                self.Warning_102_log(warning_packt)
                                
                            else:
                                temp_Rack_load_list.append(dat[1:])
                        
                    else:
                        #self.update_to_sql_Input_log(dat[1:])
                        #self.update_to_rack(dat[1:])
                        # remove input duplacated data
                        if stack_load_list !=[]:
                            if dat[1] in stack_load_list[:][0]:
                                res = next(x for x, val in enumerate(stack_load_list[:][0]) if val == dat[1]) 
                                #duplicate_load_list.append(stack_load_list[res])
                                temp_Rack_load_list.append(stack_load_list[res])
                        
                        if duplicated_prod:
                            #self.delete_product_Rack([dat[2]])
                            duplicated_in_rack.append(dat[2])


                        stack_load_list.append(dat[1:])
                        #self.update_to_sql_Rack(dat[1:])

                # if unload Type
                elif dat[0] == 1:
                    # Unload
                    available  = self.check_if_product_available(dat[2])
                    temp_prod_avail = self.check_product_available_in_temp_rack(dat[2])
                        
                    if available and temp_prod_avail:

                        Unload_info = self.get_data_from_sql_Rack_by_product(
                            dat[2])
                        
                        Unload_info_list = list(Unload_info[0])
                        Unload_info_list[22] = "Unloaded"
                        now = datetime.datetime.now()
                        current_time = now.strftime(r"%m-%d-%Y %H:%M:%S")
                        Unload_info_list[23] = current_time
                        stack_unload_list.append(Unload_info_list)
                        temp_Rack_unload_list.append(dat[2])
                        
                    elif available and (not temp_prod_avail):

                        Unload_info = self.get_data_from_sql_Rack_by_product(
                            dat[2])
                        
                        Unload_info_list = list(Unload_info[0])
                        Unload_info_list[22] = "Unloaded"
                        now = datetime.datetime.now()
                        current_time = now.strftime(r"%m-%d-%Y %H:%M:%S")
                        Unload_info_list[23] = current_time
                        stack_unload_list.append(Unload_info_list)
                        


                    elif (not available) and temp_prod_avail:
                        temp_Rack_unload_list.append(dat[2])

                    # if error occur raise dialog
                    else:
                        Error = True


                        self.Updating_dialog.done(0)
                        self.Error_001_log(dat[2])
                        self.Show_Input_ProductID_Notfound_dialog(dat[2])
                        return False


            # Update None infomatic data glove in main rack
            #self.Updating_dialog.done(0)
            
            Re_search_product = self.None_product
            self.None_temp_product = self.get_blank_TEMP_product_from_db_()
            Re_search_Temp_prodcut = self.None_temp_product

            rework_data = []


            # Transform Data shape
            for i in Re_search_product:
                 rework_data.append([2, i[0], i[1],i[2]])

            # Get infomation
            Rw_detailed_info= get_card_info_for_update(rework_data,Updated_dataframe)
            # Remove unupdate information



            if (Rw_detailed_info != None and Rw_detailed_info != []):
                for i in Rw_detailed_info:
                    #print(i)

                    stack_update_list.append(i[1:])
            
            # Update None infomatic data glove in temp rack
            rework_data = []
            update_to_temp_Rack_update =[]
            for i in Re_search_Temp_prodcut:
                 rework_data.append([2, i[0], i[1],i[2]])
            # Get infomation
            Rw_detailed_info= get_card_info_for_update(rework_data,Updated_dataframe)
            
            # Remove unupdate information



            if (Rw_detailed_info != None and Rw_detailed_info != []):
                for i in Rw_detailed_info:
                    #print(i)
                    update_to_temp_Rack_update.append(i[1:])




        # Put to all update to sqlite database
            self.Updating_dialog.done(0)
            # Close update dialog
            # Update None Product first
            if( Error != True and stack_update_list !=[]):
                self.update_to_rack(stack_update_list)

            if( Error != True and update_to_temp_Rack_update !=[]):    
                self.update_old_data_temp_Rack(update_to_temp_Rack_update)

            
            if( Error != True and duplicated_in_rack !=[]):    
                self.delete_product_Rack(duplicated_in_rack)

            if( Error != True and duplicated_in_temp_rack !=[]):    
                for i in duplicated_in_temp_rack:
                    self.delete_in_temp_rack(i)

            # then Unload 
            if( Error != True and stack_unload_list !=[]):
                delete_product = [i[1] for i in stack_unload_list] 
                #for i in stack_unload_list:    
                self.update_to_output_log(stack_unload_list)
                #self.delete_data_from_Rack_By_product(i[1])
                self.delete_product_Rack(delete_product)

            if (Error != True and temp_Rack_unload_list !=[]):
                for i in temp_Rack_unload_list:
                    self.delete_in_temp_rack(i)

            # then Load 
            if( Error != True and stack_load_list !=[]):
                self.update_to_Input_log(stack_load_list)
                self.update_to_rack(stack_load_list)
                self.update_to_temp_Rack(temp_Rack_load_list)

            # if everything is fine raise update finish
            if Error != True:
                self.Updating_dialog.done(0)
                self.Show_Update_complete_dialog()
                
                return True
        # if Error occur quit operation and rais error dialog  
        else:

            self.Updating_dialog.done(0)
            self.Show_data_base_error_log()
            
            return False
  #------------------------------------- End Update Data ------------------------------------------------

 # User Error Collecting
  # DETECT Non PAIR POSITION 
    def get_non_pair(self,test_lst):
        lst_len = len(test_lst)
        target = []
        index = 0
        while( index in range(lst_len)):
            if test_lst[index] == test_lst[-1]:
                target.append(test_lst[index]) 
                break
            else:
                first = test_lst[index][:-1]
                second = test_lst[index+1][:-1]
                if(first != second):
                    target.append(test_lst[index]) 
                    index = index + 1
                else:
                    index = index + 2

        return target

    def get_non_pair_df(self):
        test_lst = self.get_nonpair_blank_from_db()
        test_lst = [item for [item] in test_lst]
        test_lst.sort()
        non_pair_pos = []
        non_pair_df = pd.DataFrame()
        if test_lst is not None and test_lst != []:
            non_pair_pos = self.get_non_pair(test_lst)
            non_pair_df = pd.DataFrame(non_pair_pos,columns=['Rack ที่ไม่มีคู่'])
            
        return non_pair_df

 # User Interact Dialog
  # -------------------------------------- User Input and asking Dialog ----------------------------------------------

    def Ask_for_overwrite_load(self, To_write_Product_ID, Rack_ID, Current_Product_ID):
        Box = QMessageBox()
        Box.setIcon(QMessageBox.Critical)
        buttonReply = QMessageBox.question(Box, 'WIP Storage System', "Do you want to overwrite {} at {} with {} ?".format(
            Current_Product_ID, Rack_ID, To_write_Product_ID), QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if buttonReply == QMessageBox.Yes:
            # print("yes")
            return True
        else:
            # print("No")
            return False

    def Ask_for_confirm_reset_rack(self):
        Box = QMessageBox()
        buttonReply = QMessageBox.question(
            Box, 'WIP Storage System', "Please confirm reset rack", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if buttonReply == QMessageBox.Yes:
            # print("yes")
            return True
        else:
            # print("No")
            return False

    def Ask_for_refresh_data_rack(self):
        Box = QMessageBox()
        buttonReply = QMessageBox.question(
            Box, 'WIP Storage System', "Do you want to refresh Data ?", QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel, QMessageBox.Cancel)
        if buttonReply == QMessageBox.Yes:
            # print("yes")
            return True
        elif buttonReply == QMessageBox.No:
            # print("No")
            return False
        elif buttonReply == QMessageBox.Cancel:
            
            return None


    def Ask_for_Update_data_rack(self):
        Box = QMessageBox()
        buttonReply = QMessageBox.question(
            Box, 'WIP Storage System', "Please confirm to pull database", QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel, QMessageBox.Cancel)
        if buttonReply == QMessageBox.Yes:
            # print("yes")
            return True
        elif buttonReply == QMessageBox.No:
            # print("No")
            return None
        elif buttonReply == QMessageBox.Cancel:
            
            return None

    def Ask_for_confirm_reset_Datalog(self):
        Box = QMessageBox()
        buttonReply = QMessageBox.question(
            Box, 'WIP Storage System', "Please confirm reset Datalog", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if buttonReply == QMessageBox.Yes:
            # print("yes")
            return True
        else:
            # print("No")
            return False

    def Ask_for_confirm_reset_Input_Datalog(self):
        Box = QMessageBox()
        buttonReply = QMessageBox.question(
            Box, 'WIP Storage System', "Please confirm reset Inputlog", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if buttonReply == QMessageBox.Yes:
            # print("yes")
            return True
        else:
            # print("No")
            return False
  # -------------------------------------- End of User Input and asking Dialog ---------------------------------------

 # pop up Message dialog 
  #------------------------------------- Pop Up Part------------------------------------------------------
  # Error
   # When SAP file is not as setting
    def Show_data_base_error_log(self):
        msg1 = QMessageBox()
        msg1.setIcon(QMessageBox.Critical)
        msg1.setText(
            "Error : Wrong SAP Excel files ")
        msg1.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg1.setStandardButtons(QMessageBox.Ok)
        msg1.exec_()
   # When product is duplicated
    def Show_duplicate_value_log(self, ProductID):
        msg1 = QMessageBox()
        msg1.setIcon(QMessageBox.Critical)
        msg1.setText(
            "Error : ProductID : {} is duplicated ".format(ProductID))
        msg1.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg1.setStandardButtons(QMessageBox.Ok)
        msg1.exec_()
   # When not found Product in rack
    def Show_Input_ProductID_Notfound_dialog(self, ProductID):
        msg1 = QMessageBox()
        msg1.setIcon(QMessageBox.Critical)
        msg1.setText(
            "Error : Not found productID : {}  Please check the input file and reload.".format(ProductID))
        msg1.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg1.setStandardButtons(QMessageBox.Ok)
        msg1.exec_()
   # When wrong position Scannned
    def Show_Input_Rack_positton_notfound_dialog(self, Rack_position):
        msg1 = QMessageBox()
        msg1.setIcon(QMessageBox.Critical)
        msg1.setText(
            "Error : Not found rack position : {}".format(Rack_position))
        msg1.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg1.setStandardButtons(QMessageBox.Ok)
        msg1.exec_()
   # When file input is not as setting 
    def Show_Input_Wrong_file_dialog(self):
        msg1 = QMessageBox()
        msg1.setIcon(QMessageBox.Critical)
        msg1.setText("Error : Wrong input file ")
        msg1.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg1.setStandardButtons(QMessageBox.Ok)
        msg1.exec_()
   # Unknown Error
    def Show_Unknow_Error_dialog(self):
        msg1 = QMessageBox()
        msg1.setIcon(QMessageBox.Critical)
        msg1.setText("Unknown Error")
        msg1.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg1.setStandardButtons(QMessageBox.Ok)
        msg1.exec_()
   # When file Permission Error
    def Show_Permission_Error_dialog(self):
        msg1 = QMessageBox()
        msg1.setIcon(QMessageBox.Critical)
        msg1.setText("Permission Error")
        msg1.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg1.setStandardButtons(QMessageBox.Ok)
        msg1.exec_()
   # When No data in Barcode
    def Show_No_data_to_update_dialog(self):
        loadmsg = QMessageBox()
        loadmsg.setIcon(QMessageBox.Information)
        loadmsg.setText("No_data_to_update")
        loadmsg.setWindowTitle("WIP WAREHOUSE SYSTEM")
        loadmsg.setStandardButtons(QMessageBox.Ok)
        loadmsg.exec_()
  # Warning
   # When Path Not selected
    def Show_Please_Select_Path_dialog(self):
        msg1 = QMessageBox()
        msg1.setIcon(QMessageBox.Warning)
        msg1.setText("Please Select file path Before Process")
        msg1.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg1.setStandardButtons(QMessageBox.Ok)
        msg1.exec_()
  # Notify
   # Show to wait for program working
    def Show_Please_Wait_Update_dialog(self):
        self.Updating_dialog = QMessageBox()
        #self.Updating_dialog.setIcon(QMessageBox.Information)
        self.Updating_dialog.setText("Updating Product Data. \nPlease wait ...")
        self.Updating_dialog.setWindowTitle("WIP WAREHOUSE SYSTEM")
        #msg1.removeButton(QMessageBox.Ok)
        self.Updating_dialog.setStandardButtons(QMessageBox.NoButton)
        self.Updating_dialog.exec_()
   # When Reset Datalog complete
    def Show_Reset_Datalog_complete(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("Reset Datalog complete")
        msg.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
   # When Reset Rack complete
    def Show_Reset_Rack_complete(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("Reset Rack complete")
        msg.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
   # When reupdate product complete
    def Show_Reupdate_data_complete(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(
            "Re-Update Data Complete")
        msg.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
   # When Save Setting complete
    def Show_save_complete(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("Save setting complete")
        msg.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
   # When finished update Max digit of product
    def Show_Update_Max_digit_complete_dialog(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("Set Maximum ProductID digit complete")
        msg.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
   # When Export Complete
    def Show_Export_complete_dialog(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("Export File Complete")
        msg.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
   # When Update Complete
    def Show_Update_complete_dialog(self):
        loadmsg = QMessageBox()
        loadmsg.setIcon(QMessageBox.Information)
        loadmsg.setText("Update WIP data Complete")
        loadmsg.setWindowTitle("WIP WAREHOUSE SYSTEM")
        loadmsg.setStandardButtons(QMessageBox.Ok)
        loadmsg.exec_()
   # When finished database pull to temporary db
    def Show_Finish_Pull_data(self):
        loadmsg = QMessageBox()
        loadmsg.setIcon(QMessageBox.Information)
        loadmsg.setText("Pull Data from data base finished")
        loadmsg.setWindowTitle("WIP WAREHOUSE SYSTEM")
        loadmsg.setStandardButtons(QMessageBox.Ok)
        loadmsg.exec_()
   # When finishihed set Auten Query
    def Show_Please_set_Query_Authen_dialog(self):
        msg1 = QMessageBox()
        msg1.setIcon(QMessageBox.Information)
        msg1.setText("Set Query Path complete")
        msg1.setWindowTitle("WIP WAREHOUSE SYSTEM")
        msg1.setStandardButtons(QMessageBox.Ok)
        msg1.exec_()
  #------------------------------------- End Pop up Part ------------------------------------------------

 # Setting 
  #----------------------------------- Setting function -----------------------------------

    def set_max_number(self, max_number):
        self.max_number = max_number
        self.Show_Update_Max_digit_complete_dialog()

    def load_setting_value(self):
        # read Json file for setting value
        # Max_number
        # database path
        f = open('Software_Setting.json')
        WIP_setting = json.load(f)
        self.max_number = int(WIP_setting['ProductID']['max_number'])
        self.excel_database_path = WIP_setting['Database_path']['path']
        self.sql_server = WIP_setting['Server']['Server_path']
        self.sql_database= WIP_setting['Database']['Database']
        self.sql_username = WIP_setting['User']['Username']
        self.sql_password = WIP_setting['User']['Password']

    def set_query_user(self,server,database,username,password):
        self.sql_server = server
        self.sql_database = database
        self.sql_username = username
        self.sql_password = password
        self.Show_Please_set_Query_Authen_dialog()

    def save_setting_value(self):
        # save setting to Json file
        # Max number
        # database path
        # save
        WIP_software_Setting = {}
        WIP_software_Setting['ProductID'] = {
            'max_number': "{}".format(self.max_number)}
        WIP_software_Setting['Database_path'] = {
            'path': "{}".format(self.excel_database_path)}

        WIP_software_Setting['Server'] = {
            'Server_path': "{}".format(self.sql_server)}
        WIP_software_Setting['Database'] = {
            'Database': "{}".format(self.sql_database)}
        WIP_software_Setting['User'] = {
            'Username': "{}".format(self.sql_username) , 'Password' : "{}".format(self.sql_password)}


        with open('Software_Setting.json', 'w') as f:
            json.dump(WIP_software_Setting, f)

    def Reset_Rack_SQL(self):
        Dicision = self.Ask_for_confirm_reset_rack()
        if Dicision:
            self.reset_SQL_Rack_table_del_all_current()
            self.Show_Reset_Rack_complete()

 # Event Handling
  # ------------------------------------- Updating Handiling Task  ----------------------------
    def Update_Handler(self):
        # Check if file path is blank
        if self.InputPath != None and self.InputPath != "":
            self.updated_data = self.Load_unload_separator((self.get_data(self.InputPath)))

            # Code Below will be remoce later
            # self.product_barcode_to_query = self.updated_data_to_query(self.updated_data)
            self.None_product = self.get_blank_product_from_db()
            self.None_temp_product = self.get_blank_TEMP_product_from_db_()
            
            # Make Query String
            self.Re_update_product = self.get_None_product_to_query(self.None_product)
            self.Re_update_temp_product = self.get_None_product_to_query(self.None_temp_product)

            # Combine Query String
            
            if self.Re_update_product != "" :
               self.product_barcode_to_query = self.product_barcode_to_query + "," + self.Re_update_product
               #print(self.product_barcode_to_query)

            if self.Re_update_temp_product != "" :
                self.product_barcode_to_query = self.product_barcode_to_query + "," + self.Re_update_temp_product
                #print(self.product_barcode_to_query)

            # End removable 


            if self.updated_data != None and self.updated_data != []:

            # Start Update value Result is to check Error
                refresh_Excel_data = self.Ask_for_refresh_data_rack()
                if refresh_Excel_data:
                    

                    worker = Worker(self.Query_Data) # Any other args, kwargs are passed to the run function
                #worker.signals.result.connect(self.print_output)
                #worker.signals.progress.connect(self.progress_fn)
                    worker.signals.result.connect(self.Update_data)

                # Execute
                    self.threadpool.start(worker)
                    self.Show_Please_Wait_Update_dialog()
                    

                    
                elif refresh_Excel_data == False:

                    product_data_lookup = pd.read_pickle("backup_data_info.pkl")
                    #print(product_data_lookup)
                    self.Update_data(product_data_lookup)

                elif refresh_Excel_data == None :
                    return None 

            else:
                self.Show_Input_Wrong_file_dialog()
                self.Error_003_log(self.InputPath)
                # Result = self.Update_data()

                # if Result == True:
                #     # NO Error
                #     self.Show_Update_complete_dialog()
            
        else:
            
            self.Show_Please_Select_Path_dialog()
  # ------------------------------------  End Updating Task ----------------------------------------------------------------
  # ------------------------------------- Setting save Handling Task ----------------------------------------------
    def save_setting_Handler(self):
        self.save_setting_value()
        self.Show_save_complete()

    def Reset_datalog(self):
        if(self.Ask_for_confirm_reset_Datalog()):
            self.delete_all_record_log()
            self.Show_Reset_Datalog_complete()

    def Reset_data_input_log(self):
        if(self.Ask_for_confirm_reset_Input_Datalog()):
            self.delete_all_input_log()
            self.Show_Reset_Datalog_complete()
  # ------------------------------------- End Setting save Task -----------------------------------------
  # ------------------------------------- ReUpdate Handling task -------------------------

    def Rework_None_product(self,product_data_lookup):

        self.Updating_dialog.done(0)
        if product_data_lookup.empty :
            self.Show_data_base_error_log()
            return None
            
        else :
            #product_data_lookup.to_pickle("backup_data_info.pkl")
            self.Update_data_to_pickle(product_data_lookup)

        stack_update_list = []
        temp_Rack_load_list = []
        Re_search_product = self.None_product
        Re_search_Temp_prodcut = self.None_temp_product
        #print(Re_search_product)

        rework_data = []
        for i in Re_search_product:
            rework_data.append([2, i[0], i[1],i[2]])

        #print(rework_data)
        #product_data_lookup = pd.read_pickle("backup_data_info.pkl")
        Rw_detailed_info = get_card_info_for_update(rework_data,product_data_lookup)
        #print(Rw_detailed_info)
        if (Rw_detailed_info != None and Rw_detailed_info != []):
            for i in Rw_detailed_info:
                
                #self.update_to_sql_Rack(i[1:])
                stack_update_list.append(i[1:])

        # for temp rack
        rework_data = []
        for i in Re_search_Temp_prodcut:
                rework_data.append([2, i[0], i[1],i[2]])
        Rw_detailed_info= get_card_info_for_update(rework_data,product_data_lookup)
        if (Rw_detailed_info != None and Rw_detailed_info != []):
            for i in Rw_detailed_info:
                #print(i)
                temp_Rack_load_list.append(i[1:])
    

        if( stack_update_list !=[]):
            self.update_to_rack(stack_update_list)
        if( temp_Rack_load_list !=[]):    
            self.update_to_temp_Rack(temp_Rack_load_list)

        #self.make_sql_to_excel_log_Rack()    
        self.Show_Reupdate_data_complete()

    def rework_handler(self):
        refresh_Excel_data = self.Ask_for_refresh_data_rack()
        self.None_product = self.get_blank_product_from_db()
        self.None_temp_product = self.get_blank_TEMP_product_from_db_()
        self.product_barcode_to_query  = self.get_None_product_to_query(self.None_product)
        self.Re_update_temp_product = self.get_None_product_to_query(self.None_temp_product)
        
        if self.Re_update_temp_product != "" :
            self.product_barcode_to_query = self.product_barcode_to_query + "," + self.Re_update_temp_product
            #print(self.product_barcode_to_query)




        if self.product_barcode_to_query == "":
            self.Show_No_data_to_update_dialog()
            return None

        if refresh_Excel_data:
            
            
            
            

            worker = Worker(self.Query_Data) # Any other args, kwargs are passed to the run function
        #worker.signals.result.connect(self.print_output)
        #worker.signals.progress.connect(self.progress_fn)
            worker.signals.result.connect(self.Rework_None_product)

        # Execute
            self.threadpool.start(worker)
            self.Show_Please_Wait_Update_dialog()
            
        elif refresh_Excel_data == False:

            product_data_lookup = pd.read_pickle("backup_data_info.pkl")
            #print(product_data_lookup)
            
            self.Rework_None_product(product_data_lookup)


        elif refresh_Excel_data == None :
            return None 
  # ------------------------------------- End Reupdate Handling task -------------------------

if __name__ == "__main__":

    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainPage = QtWidgets.QMainWindow()
    ui = Ui_MainPage()
    ui.setupUi(MainPage)
    MainPage.show()
    sys.exit(app.exec_())


